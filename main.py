
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

TEST_CONTENTS_COLUMN = 'B'

def main() -> None:
    try:
        arg = sys.argv[1]
    except IndexError:
        arg = ''
    
    if arg == 'createFiles':
        ws = getWorksheet()
        createTestFiles(ws)

    exit(0)


def getWorksheet() -> Worksheet:
    wb = load_workbook('Testing.xlsx', data_only=True, read_only=True)
    sheetName = wb.sheetnames[0]
    ws = wb[sheetName]
    return ws


def createTestFiles(ws: Worksheet):

    try:
        os.makedirs('Tests')
    except:
        for file in os.listdir('Tests/'):
            os.remove('Tests/' + file)

    testCount = 0

    for i,row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        
        testNo = i
        testContent = row[1].value

        if testContent is None:
            break

        testName = 'Tests/' + str(testNo) + '.txt'

        with open(testName, 'w') as f:
            f.write(testContent)
            testCount += 1

    print(f'Successfully created {testCount} text files')


if __name__ == '__main__':
    main()